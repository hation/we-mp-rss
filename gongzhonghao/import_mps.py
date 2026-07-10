#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
功能1：从 Excel 批量导入公众号到 WeRSS。

用法:
  python gongzhonghao/import_mps.py                    # 导入全部
  python gongzhonghao/import_mps.py --limit 3          # 只导入前3个（测试用）
  python gongzhonghao/import_mps.py --offset 5         # 跳过前5个
  python gongzhonghao/import_mps.py --file other.xlsx  # 指定其他Excel

Excel格式：第一个sheet，第一列为公众号名称，带表头跳过。
"""
import argparse
import time
import urllib.parse

import openpyxl
import requests

from _common import BASE, XLSX_PATH, login, auth_headers, print_summary

SEARCH_INTERVAL = 2  # 搜索间隔（秒），避免触发微信风控
ADD_INTERVAL = 3     # 添加间隔（秒）


def load_names(path: str) -> list[str]:
    """从Excel第一列读取公众号名称，跳过空行和表头。"""
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    names = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not row or not row[0]:
            continue
        val = str(row[0]).strip()
        if not val or i == 0:  # 跳过表头
            continue
        names.append(val)
    return names


def search_mp(token: str, name: str) -> dict | None:
    """搜索公众号，优先返回 nickname 完全匹配项，否则返回第一项。"""
    url = f"{BASE}/wx/mps/search/{urllib.parse.quote(name)}"
    r = requests.get(url, headers=auth_headers(token), timeout=30)
    r.raise_for_status()
    items = r.json().get("data", {}).get("list", [])
    if not items:
        return None
    for it in items:
        if it.get("nickname") == name:
            return it
    return items[0]


def add_mp(token: str, item: dict) -> dict:
    payload = {
        "mp_name": item.get("nickname", ""),
        "mp_id": item.get("fakeid", ""),
        "avatar": item.get("round_head_img", ""),
        "mp_intro": item.get("signature", "") or "",
    }
    r = requests.post(
        f"{BASE}/wx/mps",
        headers={**auth_headers(token), "Content-Type": "application/json"},
        json=payload,
        timeout=60,
    )
    try:
        return r.json()
    except Exception:
        return {"raw": r.text, "status": r.status_code}


def main():
    ap = argparse.ArgumentParser(description="从Excel批量导入公众号")
    ap.add_argument("--file", default=XLSX_PATH, help=f"Excel路径（默认: {XLSX_PATH}）")
    ap.add_argument("--limit", type=int, default=0, help="只导入前N个，0=全部")
    ap.add_argument("--offset", type=int, default=0, help="跳过前N个")
    args = ap.parse_args()

    names = load_names(args.file)
    if args.offset:
        names = names[args.offset:]
    if args.limit:
        names = names[: args.limit]

    print(f"待导入 {len(names)} 个公众号")
    print("=" * 60)

    token = login()
    print("登录成功")

    ok, fail, skip = [], [], []
    for i, name in enumerate(names, 1):
        print(f"\n[{i}/{len(names)}] {name}")
        try:
            time.sleep(SEARCH_INTERVAL)
            item = search_mp(token, name)
        except Exception as e:
            print(f"  搜索异常: {e}")
            fail.append((name, f"搜索异常: {e}"))
            continue

        if not item:
            print("  搜索无结果，跳过")
            skip.append((name, "搜索无结果"))
            continue

        nick = item.get("nickname", "")
        if nick != name:
            print(f"  注意: 搜索结果 '{nick}' 与目标不匹配，仍尝试添加")

        try:
            time.sleep(ADD_INTERVAL)
            resp = add_mp(token, item)
        except Exception as e:
            print(f"  添加异常: {e}")
            fail.append((name, f"添加异常: {e}"))
            continue

        if resp.get("code") == 0:
            data = resp.get("data", {})
            print(f"  成功: id={data.get('id')} mp_name={data.get('mp_name')}")
            ok.append(name)
        else:
            print(f"  失败: {resp.get('message', resp)}")
            fail.append((name, resp.get("message", "unknown")))

    print_summary("导入完成", ok, fail, skip)


if __name__ == "__main__":
    main()
